import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # creating a chart:
    values = Reference(sheet,
                        min_row = 2,
                        max_row = sheet.max_row,
                        min_col = 4,
                        max_col = 4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    new_filename = "corrected " + filename
    wb.save(new_filename)


File_Name = 'transactions.xlsx'
process_workbook(File_Name)
